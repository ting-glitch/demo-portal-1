import os
import re
import json
import datetime
import traceback
import logging
from typing import Dict, List, Optional
from pydantic import BaseModel
from fastapi import FastAPI, HTTPException, Body
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import uvicorn
import tempfile
import shutil

# We import pywin32 components inside try-except to avoid load-time failure if not registered yet
try:
    import pythoncom
    import win32com.client
    HAS_PYWIN32 = True
except ImportError:
    HAS_PYWIN32 = False

FORCE_DEMO = True

def check_outlook_available() -> bool:
    # Safe demo fallback: return False instantly if in demo mode to prevent COM hanging on PCs without Outlook configured
    if FORCE_DEMO:
        return False
    if not HAS_PYWIN32:
        return False
    try:
        pythoncom.CoInitialize()
        try:
            # CRITICAL SECURITY FIX FOR INFINITE HANG:
            # Do NOT use win32com.client.Dispatch("Outlook.Application") directly here because on machines
            # without an Outlook profile or configured account, it starts a background Outlook process
            # that displays a modal "Welcome/Login" dialog in session 0 or in the background, hanging indefinitely.
            # Instead, we use GetObject which ONLY connects to an already active Outlook instance, failing instantly
            # if Outlook is not open, which guarantees zero hanging.
            try:
                outlook = win32com.client.GetObject(Class="Outlook.Application")
            except Exception:
                return False
                
            ns = outlook.GetNamespace("MAPI")
            # Try a quick call to ensure Outlook is configured and logged in
            _ = ns.GetDefaultFolder(6)
            return True
        except Exception:
            return False
        finally:
            pythoncom.CoUninitialize()
    except Exception:
        return False

app = FastAPI(title="Logistics Operations Platform API", version="1.0.0")

# Setup paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(BASE_DIR)
DATA_FILE = os.path.join(BASE_DIR, "data.json")
TEMPLATES_DIR = os.path.join(ROOT_DIR, "frontend", "templates")

# Default configurations
DEFAULT_DOWNLOAD_DIR = os.path.join(BASE_DIR, "downloads")
LOG_FILE = os.path.join(BASE_DIR, "operation_platform.log")

# Ensure base directories exist
os.makedirs(TEMPLATES_DIR, exist_ok=True)
os.makedirs(DEFAULT_DOWNLOAD_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    filename=LOG_FILE,
    filemode="a",
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
ship_logger = logging.getLogger("operation_platform")

# Define Schemas for Extensibility (Phases 1, 2, and 3)
class Phase1Status(BaseModel):
    completed: bool = False
    attachment_saved: bool = False
    forward_prepared: bool = False
    completed_time: Optional[str] = None

class Phase2Status(BaseModel):
    completed: bool = False
    available: bool = False
    shipped: bool = False
    completed_time: Optional[str] = None
    ba_no: str = ""
    entry_id: str = ""
    data: Dict = {}

class Phase3Status(BaseModel):
    completed: bool = False
    completed_time: Optional[str] = None
    data: Dict = {}

class Phases(BaseModel):
    phase1: Phase1Status = Phase1Status()
    phase2: Phase2Status = Phase2Status()
    phase3: Phase3Status = Phase3Status()

class Record(BaseModel):
    id: str  # Unique ID
    entry_id: str  # Original Outlook EntryID
    center: str = "EMS-A"  # "Center" (originally RTR)
    type: str  # "OK" or "BAD"
    ship_out_date: str  # YYYY/MM/DD
    qty: int
    creation_date: str = ""  # Creation Date (originally bpm_date)
    er: str = ""  # ER (originally ems_no / bpm_no)
    rma_no: str = ""  # Only for BAD type
    ba_no: str = ""  # Extracted from RPA reply, displayed as Phase 2 identifier
    return_to: str = ""
    status: str = "New"  # "New", "Processing", "RPA Submitted", etc.
    current_phase: int = 1
    phases: Phases = Phases()
    subject: str
    received_time: str
    attachments: List[str] = []
    keyword: str = ""  # Extracted from packing list filename
    batch: str = ""  # Extracted from packing list filename
    transfer_type: str = ""  # Calculated
    calculated_subject: str = ""  # Computed forwarding subject

class Config(BaseModel):
    download_dir: str = DEFAULT_DOWNLOAD_DIR

class Database(BaseModel):
    config: Config = Config()
    records: List[Record] = []

# Helper functions to load/save database
def load_db() -> Database:
    if not os.path.exists(DATA_FILE):
        db = Database()
        save_db(db)
        return db
    try:
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            return Database.model_validate(data)
    except Exception as e:
        print(f"Error loading database: {e}")
        return Database()

def save_db(db: Database):
    try:
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(db.model_dump(), f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"Error saving database: {e}")

# Helper functions for Parsing
def parse_email_subject(subject: str) -> Optional[dict]:
    pattern = r"\[APAC\]\s*Repair\s*(好品|坏品|壞品|良品|不良品)\s*_\s*入\s*Hub-G\s*_\s*(\d{6}|\d{8})\s*_\s*(\d+)\s*PCS"
    match = re.search(pattern, subject.strip(), re.IGNORECASE)
    if not match:
        return None
    
    type_str = match.group(1)
    date_str = match.group(2)
    qty_str = match.group(3)
    
    # Determine type (OK or BAD)
    if any(x in type_str for x in ["好", "良"]):
        parsed_type = "OK"
    else:
        parsed_type = "BAD"
        
    # Format Date nicely
    formatted_date = date_str
    if len(date_str) == 6:
        formatted_date = f"20{date_str[0:2]}/{date_str[2:4]}/{date_str[4:6]}"
    elif len(date_str) == 8:
        formatted_date = f"{date_str[0:4]}/{date_str[4:6]}/{date_str[6:8]}"
        
    return {
        "center": "Hub",
        "type": parsed_type,
        "ship_out_date": formatted_date,
        "qty": int(qty_str),
        "return_to": "Hub-G" if parsed_type == "OK" else "Hub-P"
    }

def extract_keyword_from_attachment(filename: str) -> str:
    fn = filename.strip()
    prefix = "packing list-"
    if fn.lower().startswith(prefix):
        fn = fn[len(prefix):]
    else:
        return ""
        
    if "." in fn:
        fn = fn.rsplit(".", 1)[0]
        
    for suffix in ["良品", "不良品", "好品", "壞品", "坏品", "ok", "bad", "OK", "BAD"]:
        if fn.endswith(suffix):
            fn = fn[:-len(suffix)]
            break
            
    return fn.strip()

def extract_batch_from_attachments(filenames: List[str]) -> str:
    prefix = "packing list-"
    priorities = ["HUB", "EU", "WEST", "DOA"]
    batch_values = []

    for priority in priorities:
        found_values = []
        for filename in filenames:
            fn = filename.strip()
            if not fn.lower().startswith(prefix):
                continue
            core = fn[len(prefix):]
            if "." in core:
                core = core.rsplit(".", 1)[0]

            pattern = rf"({priority}(?:-[^\u4e00-\u9fff\s]+)?)"
            match = re.search(pattern, core, re.IGNORECASE)
            if match:
                value = match.group(1)
                normalized = value.upper() if priority != "EU" else value
                if normalized not in found_values:
                    found_values.append(normalized)

        if found_values:
            batch_values.extend(found_values)

    return "&".join(batch_values)

def export_excel():
    db = load_db()
    wb = Workbook()
    
    # Define styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10, color="000000")
    
    summary_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    mapping_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    # ----------------- Sheet 1: Records_Summary -----------------
    ws1 = wb.active
    ws1.title = "Records_Summary"
    
    headers1 = [
        "ID", "Center", "Batch", "Type", "Ship-out Date", "Creation Date", "Tracking No", "RMA No",
        "Qty (PCS)", "Return to", "Phase 1 Time", "Status", "Keyword",
        "Transfer Type", "Forward Subject"
    ]
    ws1.append(headers1)
    
    for r in db.records:
        ws1.append([
            r.id,
            r.center,
            r.batch or "",
            r.type,
            r.ship_out_date,
            r.creation_date or "",
            r.er or "",
            r.rma_no or "",
            r.qty,
            r.return_to or "",
            r.phases.phase1.completed_time or "",
            r.status,
            r.keyword or "",
            r.transfer_type or "",
            r.calculated_subject or ""
        ])
        
    for col_idx in range(1, len(headers1) + 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = summary_fill
        cell.alignment = align_center
        cell.border = border_all
    
    for row_idx in range(2, ws1.max_row + 1):
        for col_idx in range(1, len(headers1) + 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = border_all
            if col_idx in [13, 15]:
                cell.alignment = align_left
            else:
                cell.alignment = align_center
                
    for col in ws1.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            val_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
            if val_len > max_len:
                max_len = val_len
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # ----------------- Sheet 2: BPM_RMA_Mapping -----------------
    ws2 = wb.create_sheet(title="ER_RMA_Mapping")
    
    headers2 = [
        "Tracking No", "RMA No", "Type", "Ship-out Date", "Qty (PCS)", "Status"
    ]
    ws2.append(headers2)
    
    for r in db.records:
        if r.er:
            if r.type == "BAD" and r.rma_no:
                rmas = [x.strip() for x in r.rma_no.split(",") if x.strip()]
                for rma in rmas:
                    ws2.append([
                        r.er,
                        rma,
                        r.type,
                        r.ship_out_date,
                        r.qty,
                        r.status
                    ])
            else:
                ws2.append([
                    r.er,
                    r.rma_no or "--",
                    r.type,
                    r.ship_out_date,
                    r.qty,
                    r.status
                ])
                
    for col_idx in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = mapping_fill
        cell.alignment = align_center
        cell.border = border_all
        
    for row_idx in range(2, ws2.max_row + 1):
        for col_idx in range(1, len(headers2) + 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = border_all
            cell.alignment = align_center
            
    for col in ws2.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            val_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
            if val_len > max_len:
                max_len = val_len
        ws2.column_dimensions[col_letter].width = max(max_len + 4, 15)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"Operation_records_{datetime.date.today().strftime('202X%m%d')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

def extract_phase2_info(subject: str) -> Optional[dict]:
    # Subject format: Repair_{TYPE}_return to hub_{TRANSFER_TYPE}_{TARGET}_{YYYYMMDD}_BA-*
    pattern = r"Repair_(OK|BAD)_return to hub_(\w+)_(\w+)_(\d{8})_(?:BA|TRK-XX)-([\w-]+)"
    match = re.search(pattern, subject.strip(), re.IGNORECASE)
    if not match:
        # Fallback to old format just in case
        pattern_old = r"Repair_(OK|BAD)_return to hub_(\w+) to (\w+)_(\w+)_(\d{8})_(?:BA|TRK-XX)-([\w-]+)"
        match_old = re.search(pattern_old, subject.strip(), re.IGNORECASE)
        if not match_old:
            return None
        return {
            "type": match_old.group(1).upper(),
            "source": match_old.group(2).lower(),
            "date": match_old.group(5),
            "ba_no": match_old.group(6),
            "transfer_type": match_old.group(2).lower()
        }
    
    record_type = match.group(1).upper()
    transfer_type = match.group(2).lower()
    target = match.group(3)
    date_str = match.group(4)
    ba_no = match.group(5)

    return {
        "type": record_type,
        "source": transfer_type,
        "date": date_str,
        "ba_no": ba_no,
        "transfer_type": transfer_type
    }

def extract_phase3_info(subject: str) -> Optional[dict]:
    pattern = r"Repair_BAD_return to hub_.*_(?:BA|TRK-XX)-([\w-]+)"
    match = re.search(pattern, subject.strip(), re.IGNORECASE)
    if not match:
        return None
    return {
        "ba_no": match.group(1)
    }

def get_phase1_completed_date(record: Record) -> Optional[str]:
    if not record.phases.phase1.completed_time:
        return None
    match = re.match(r"(\d{4})/(\d{2})/(\d{2})", record.phases.phase1.completed_time)
    if not match:
        return None
    return f"{match.group(1)}{match.group(2)}{match.group(3)}"

def record_matches_phase2(record: Record, info: dict) -> bool:
    if record.type != info["type"]:
        return False
    
    batch_lower = record.batch.lower()
    if info["source"] == "virtual":
        if not ("hub" in batch_lower or "eu" in batch_lower): return False
    elif info["source"] == "real":
        if not ("west" in batch_lower or "doa" in batch_lower): return False
    else:
        return False

    phase1_date = get_phase1_completed_date(record)
    if phase1_date == info["date"]:
        return True
        
    ship_date_norm = record.ship_out_date.replace("/", "")
    if ship_date_norm == info["date"]:
        return True
    
    return False

def find_phase2_reply_for_record(record: Record, ns) -> Optional[dict]:
    try:
        inbox = ns.GetDefaultFolder(6)
        messages = inbox.Items
        messages.Sort("[ReceivedTime]", True)
        limit = min(120, messages.Count)
        ship_logger.info("Searching inbox for phase2 reply for record_id=%s limit=%s", record.id, limit)
        for i in range(1, limit + 1):
            try:
                msg = messages.Item(i)
                subject = str(msg.Subject or "")
                info = extract_phase2_info(subject)
                if not info:
                    continue
                if record_matches_phase2(record, info):
                    ship_logger.info("Found phase2 reply in inbox: %s", subject)
                    return {
                        "info": info,
                        "entry_id": str(msg.EntryID),
                        "received_time": getattr(msg, "ReceivedTime", None)
                    }
            except Exception as ex:
                ship_logger.info("Ignoring inbox message during phase2 search: %s", str(ex))
                continue
    except Exception as ex:
        ship_logger.info("Phase2 reply search failed: %s", str(ex))
    return None

def calculate_transfer_and_subject(record_type: str, keyword: str, today_str: str) -> dict:
    kw_lower = keyword.lower()
    source = "virtual"  # default fallback
    
    if "west" in kw_lower or "doa" in kw_lower:
        source = "real"
    elif "hub" in kw_lower or "eu" in kw_lower:
        source = "virtual"
        
    transfer_type = source  # "virtual" or "real"
    
    if record_type == "BAD":
        target = "Hub"
    else:
        target = "Hub-G"
        
    calculated_subject = f"Repair_{record_type}_return to hub_{transfer_type}_{target}_{today_str}"
    
    return {
        "transfer_type": transfer_type,
        "calculated_subject": calculated_subject
    }

def get_week_number(date_obj: datetime.date) -> str:
    year, week_num, weekday = date_obj.isocalendar()
    return f"W{week_num:02d}"

ATTACHMENT_TYPE_KEYWORDS = {
    "packing list": ["packing list", "packing_list", "packinglist"],
    "shipping list": ["shipping invoice", "shipping list", "不良品"]
}

def has_type(file_paths: List[str], attachment_type: str) -> bool:
    keywords = ATTACHMENT_TYPE_KEYWORDS.get(attachment_type, [attachment_type])
    for path in file_paths:
        name_lower = os.path.basename(path).lower()
        for keyword in keywords:
            if keyword in name_lower:
                return True
    return False

# Simulated email draft generators (Demo Mode HTML Compose windows)
def generate_simulated_outlook_window(subject: str, to_address: str, cc_address: str, body: str, attachments: List[str], filename_prefix: str) -> str:
    safe_filename = f"{filename_prefix}_{datetime.datetime.now().strftime('202X%m%d_%H%M%S')}.html"
    downloads_dir = os.path.join(BASE_DIR, "downloads")
    os.makedirs(downloads_dir, exist_ok=True)
    filepath = os.path.join(downloads_dir, safe_filename)
    
    attachment_items_html = "".join([f'<span style="display:inline-block; background:#f3f2f1; border:1px solid #d2d0ce; border-radius:4px; padding:4px 8px; margin:4px; font-size:10pt; font-family:sans-serif; color:#323130;"><span style="color:#107c41; font-weight:bold; margin-right:4px;">📊</span> {name}</span>' for name in attachments])
    
    html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Simulated Outlook Draft</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background-color: #f3f2f1;
            margin: 0;
            padding: 20px;
            color: #323130;
        }}
        .window {{
            max-width: 800px;
            margin: 0 auto;
            background: #ffffff;
            border: 1px solid #e1dfdd;
            box-shadow: 0 4px 16px rgba(0, 0, 0, 0.15);
            border-radius: 4px;
            overflow: hidden;
        }}
        .header {{
            background: #0078d4;
            color: #ffffff;
            padding: 12px 16px;
            display: flex;
            align-items: center;
            justify-content: space-between;
        }}
        .header-title {{
            font-weight: 600;
            font-size: 11pt;
            display: flex;
            align-items: center;
        }}
        .header-logo {{
            background: #ffffff;
            color: #0078d4;
            width: 20px;
            height: 20px;
            border-radius: 2px;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            font-weight: bold;
            font-size: 10pt;
            margin-right: 8px;
        }}
        .demo-badge {{
            background: #fff176;
            color: #333333;
            font-size: 8pt;
            font-weight: bold;
            padding: 2px 6px;
            border-radius: 4px;
            text-transform: uppercase;
        }}
        .fields-table {{
            width: 100%;
            border-collapse: collapse;
            border-bottom: 1px solid #f3f2f1;
        }}
        .fields-table td {{
            padding: 10px 16px;
            border-bottom: 1px solid #f3f2f1;
            vertical-align: top;
            font-size: 10pt;
        }}
        .label {{
            color: #605e5c;
            width: 80px;
            font-weight: 600;
        }}
        .value {{
            color: #323130;
        }}
        .subject-val {{
            font-weight: 600;
            color: #201f1e;
        }}
        .attachments-container {{
            background: #faf9f8;
            padding: 8px 16px;
            border-bottom: 1px solid #e1dfdd;
        }}
        .attachments-title {{
            font-size: 9pt;
            font-weight: 600;
            color: #605e5c;
            margin-bottom: 6px;
        }}
        .body-area {{
            padding: 24px;
            min-height: 250px;
            font-size: 11pt;
            font-family: 'Calibri', 'Segoe UI', sans-serif;
            line-height: 1.5;
            color: #000000;
        }}
        .footer {{
            background: #faf9f8;
            border-top: 1px solid #f3f2f1;
            padding: 12px 16px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}
        .send-btn {{
            background: #0078d4;
            color: #ffffff;
            border: none;
            padding: 8px 20px;
            border-radius: 2px;
            font-weight: 600;
            font-size: 10pt;
            cursor: not-allowed;
        }}
        .status-text {{
            font-size: 9pt;
            color: #a19f9d;
            font-style: italic;
        }}
    </style>
</head>
<body>
    <div class="window">
        <div class="header">
            <span class="header-title">
                <span class="header-logo">O</span>
                Outlook - Simulated Email Draft (De-identified)
            </span>
            <span class="demo-badge">Portfolio Demo Mode</span>
        </div>
        <table class="fields-table">
            <tr>
                <td class="label">To:</td>
                <td class="value">{to_address}</td>
            </tr>
            <tr>
                <td class="label">Cc:</td>
                <td class="value">{cc_address}</td>
            </tr>
            <tr>
                <td class="label">Subject:</td>
                <td class="value subject-val">{subject}</td>
            </tr>
        </table>
        {"<div class='attachments-container'><div class='attachments-title'>Attachments ({len(attachments)})</div>" + attachment_items_html + "</div>" if attachments else ""}
        <div class="body-area">
            {body}
        </div>
        <div class="footer">
            <button class="send-btn" disabled>Send</button>
            <span class="status-text">Draft successfully simulated & displayed offline</span>
        </div>
    </div>
</body>
</html>
"""
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(html_content)
        
    import webbrowser
    file_url = "file:///" + filepath.replace("\\", "/")
    webbrowser.open(file_url)
    ship_logger.info(f"Demo Mode: Opened simulated draft in browser: {filepath}")
    return filepath

def create_mock_attachment_files(dest_dir: str, filenames: List[str]):
    os.makedirs(dest_dir, exist_ok=True)
    for fname in filenames:
        fpath = os.path.join(dest_dir, fname)
        if not os.path.exists(fpath):
            wb = Workbook()
            ws = wb.active
            ws.title = "Mock Data"
            ws.append(["Item ID", "Part Number", "Description", "Quantity (PCS)", "Status"])
            ws.append([1, "PN-998877-A1", "Standard Receiver Module", 100, "Approved"])
            ws.append([2, "PN-554433-B2", "Core Control Board", 50, "Approved"])
            wb.save(fpath)
            ship_logger.info(f"Demo Mode: Created mock Excel attachment file at {fpath}")

# COM Thread Safe Helpers
class OutlookCOMContext:
    def __enter__(self):
        if not HAS_PYWIN32:
            raise HTTPException(status_code=500, detail="pywin32 library is not installed or registered on this Windows machine.")
        pythoncom.CoInitialize()
        try:
            self.outlook = win32com.client.Dispatch("Outlook.Application")
            return self.outlook
        except Exception as e:
            pythoncom.CoUninitialize()
            raise HTTPException(status_code=500, detail=f"Failed to connect to Microsoft Outlook client: {str(e)}")

    def __exit__(self, exc_type, exc_val, exc_tb):
        pythoncom.CoUninitialize()

# Backend Routes
@app.get("/api/config", response_model=Config)
def get_config():
    db = load_db()
    return db.config

@app.post("/api/config", response_model=Config)
def update_config(config: Config):
    db = load_db()
    config.download_dir = os.path.abspath(config.download_dir.replace("/", "\\"))
    db.config = config
    save_db(db)
    return db.config

@app.get("/api/records", response_model=List[Record])
def get_records():
    db = load_db()
    return db.records

@app.delete("/api/records")
def delete_all_records():
    # portfolio demo-friendly reset: wipe and immediately re-seed 11 high-fidelity records
    db = Database()
    save_db(db) # clear
    # Now scan will auto-seed because len(records) == 0
    scan_outlook() 
    return {"status": "ok"}

@app.post("/api/records/scan", response_model=List[Record])
def scan_outlook():
    db = load_db()
    
    # We always use the robust simulation mode if Outlook is unavailable or explicitly fake (for high-fidelity step 1-3 demo)
    if not check_outlook_available() or True:
        ship_logger.info("Running scan_outlook in [HIGH-FIDELITY DEMO MODE] simulation")
        if len(db.records) == 0:
            today_str = datetime.date.today().strftime("202X/%m/%d")
            today_nodash = datetime.date.today().strftime("202X%m%d")
            
            # Record 1: OK - Shipped
            rec1 = Record(
                id="DEMO_REC_1",
                entry_id="ENTRY_1",
                center="EMS-A",
                type="OK",
                ship_out_date=(datetime.date.today() - datetime.timedelta(days=10)).strftime("202X/%m/%d"),
                qty=120,
                creation_date=(datetime.date.today() - datetime.timedelta(days=9)).strftime("202X/%m/%d"),
                er="TRK-XX-11001",
                subject="[APAC] Repair 好品_入Hub-G_20260512_120PCS",
                received_time="202X/05/12 09:30:--",
                attachments=["packing list-HUB-501.xlsx", "shipping invoice-HUB-501.xlsx"],
                keyword="HUB-501",
                batch="HUB-501",
                transfer_type="virtual",
                calculated_subject="Repair_OK_return to hub_virtual_Hub-G_20260512",
                status="Shipped"
            )
            rec1.phases.phase1.completed = True
            rec1.phases.phase1.attachment_saved = True
            rec1.phases.phase1.forward_prepared = True
            rec1.phases.phase1.completed_time = "202X/05/12 11:15:--"
            rec1.phases.phase2.completed = True
            rec1.phases.phase2.available = True
            rec1.phases.phase2.shipped = True
            rec1.phases.phase2.ba_no = "TRK-XX-11001"
            rec1.phases.phase2.completed_time = "202X/05/13 14:20:--"

            # Record 2: BAD - RMA Approved (Phase 3 Completed)
            rec2 = Record(
                id="DEMO_REC_2",
                entry_id="ENTRY_2",
                center="EMS-B",
                type="BAD",
                ship_out_date=(datetime.date.today() - datetime.timedelta(days=8)).strftime("202X/%m/%d"),
                qty=35,
                creation_date=(datetime.date.today() - datetime.timedelta(days=7)).strftime("202X/%m/%d"),
                er="TRK-XX-11002",
                rma_no="RMA-XXX-XXXX, RMA-YYY-YYYY",
                subject="[APAC] Repair 坏品_入Hub-G_20260514_35PCS",
                received_time="202X/05/14 10:15:--",
                attachments=["packing list-HUB-502.xlsx", "shipping invoice-HUB-502.xlsx", "shipping list-35PCS不良品.xlsx"],
                keyword="HUB-502",
                batch="HUB-502",
                transfer_type="virtual",
                calculated_subject="Repair_BAD_return to hub_virtual_Hub_20260514",
                status="Shipped"
            )
            rec2.phases.phase1.completed = True
            rec2.phases.phase1.attachment_saved = True
            rec2.phases.phase1.forward_prepared = True
            rec2.phases.phase1.completed_time = "202X/05/14 12:00:--"
            rec2.phases.phase2.completed = True
            rec2.phases.phase2.available = True
            rec2.phases.phase2.shipped = True
            rec2.phases.phase2.ba_no = "TRK-XX-11002"
            rec2.phases.phase2.completed_time = "202X/05/15 16:30:--"
            rec2.phases.phase3.completed = True
            rec2.phases.phase3.completed_time = "202X/05/16 10:05:--"

            # Record 3: OK - Phase 2 Ready
            rec3 = Record(
                id="DEMO_REC_3",
                entry_id="ENTRY_3",
                center="EMS-C",
                type="OK",
                ship_out_date=(datetime.date.today() - datetime.timedelta(days=5)).strftime("202X/%m/%d"),
                qty=210,
                creation_date=(datetime.date.today() - datetime.timedelta(days=4)).strftime("202X/%m/%d"),
                er="TRK-XX-11003",
                subject="[APAC] Repair 好品_入Hub-G_20260517_210PCS",
                received_time="202X/05/17 08:45:--",
                attachments=["packing list-HUB-503.xlsx", "shipping invoice-HUB-503.xlsx"],
                keyword="HUB-503",
                batch="HUB-503",
                transfer_type="virtual",
                calculated_subject="Repair_OK_return to hub_virtual_Hub-G_20260517",
                status="Phase 2 Ready"
            )
            rec3.phases.phase1.completed = True
            rec3.phases.phase1.attachment_saved = True
            rec3.phases.phase1.forward_prepared = True
            rec3.phases.phase1.completed_time = "202X/05/17 10:20:--"
            rec3.phases.phase2.available = True
            rec3.phases.phase2.ba_no = "TRK-XX-11003"

            # Record 4: BAD - Phase 2 Ready
            rec4 = Record(
                id="DEMO_REC_4",
                entry_id="ENTRY_4",
                center="EMS-A",
                type="BAD",
                ship_out_date=(datetime.date.today() - datetime.timedelta(days=4)).strftime("202X/%m/%d"),
                qty=50,
                creation_date=(datetime.date.today() - datetime.timedelta(days=3)).strftime("202X/%m/%d"),
                er="TRK-XX-11004",
                subject="[APAC] Repair 坏品_入Hub-G_20260518_50PCS",
                received_time="202X/05/18 11:20:--",
                attachments=["packing list-HUB-504.xlsx", "shipping invoice-HUB-504.xlsx", "shipping list-50PCS不良品.xlsx"],
                keyword="HUB-504",
                batch="HUB-504",
                transfer_type="virtual",
                calculated_subject="Repair_BAD_return to hub_virtual_Hub_20260518",
                status="Phase 2 Ready"
            )
            rec4.phases.phase1.completed = True
            rec4.phases.phase1.attachment_saved = True
            rec4.phases.phase1.forward_prepared = True
            rec4.phases.phase1.completed_time = "202X/05/18 13:40:--"
            rec4.phases.phase2.available = True
            rec4.phases.phase2.ba_no = "TRK-XX-11004"

            # Record 5: OK - RPA Submitted
            rec5 = Record(
                id="DEMO_REC_5",
                entry_id="ENTRY_5",
                center="EMS-B",
                type="OK",
                ship_out_date=(datetime.date.today() - datetime.timedelta(days=2)).strftime("202X/%m/%d"),
                qty=180,
                subject="[APAC] Repair 好品_入Hub-G_20260520_180PCS",
                received_time="202X/05/20 14:10:--",
                attachments=["packing list-HUB-505.xlsx", "shipping invoice-HUB-505.xlsx"],
                keyword="HUB-505",
                batch="HUB-505",
                transfer_type="virtual",
                calculated_subject="Repair_OK_return to hub_virtual_Hub-G_20260520",
                status="RPA Submitted"
            )
            rec5.phases.phase1.completed = True
            rec5.phases.phase1.attachment_saved = True
            rec5.phases.phase1.forward_prepared = True
            rec5.phases.phase1.completed_time = "202X/05/20 15:45:--"

            # Record 6: BAD - RPA Submitted
            rec6 = Record(
                id="DEMO_REC_6",
                entry_id="ENTRY_6",
                center="EMS-C",
                type="BAD",
                ship_out_date=(datetime.date.today() - datetime.timedelta(days=2)).strftime("202X/%m/%d"),
                qty=18,
                subject="[APAC] Repair 坏品_入Hub-G_20260520_18PCS",
                received_time="202X/05/20 14:30:--",
                attachments=["packing list-HUB-506.xlsx", "shipping invoice-HUB-506.xlsx", "shipping list-18PCS不良品.xlsx"],
                keyword="HUB-506",
                batch="HUB-506",
                transfer_type="virtual",
                calculated_subject="Repair_BAD_return to hub_virtual_Hub_20260520",
                status="RPA Submitted"
            )
            rec6.phases.phase1.completed = True
            rec6.phases.phase1.attachment_saved = True
            rec6.phases.phase1.forward_prepared = True
            rec6.phases.phase1.completed_time = "202X/05/20 16:15:--"

            # Record 7: OK - Processing (Attachment Saved)
            rec7 = Record(
                id="DEMO_REC_7",
                entry_id="ENTRY_7",
                center="EMS-A",
                type="OK",
                ship_out_date=(datetime.date.today() - datetime.timedelta(days=1)).strftime("202X/%m/%d"),
                qty=95,
                subject="[APAC] Repair 好品_入Hub-G_20260521_95PCS",
                received_time="202X/05/21 09:10:--",
                attachments=["packing list-WEST-701.xlsx", "shipping invoice-WEST-701.xlsx"],
                keyword="WEST-701",
                batch="WEST-701",
                transfer_type="real",
                calculated_subject="Repair_OK_return to hub_real_Hub-G_20260521",
                status="Processing"
            )
            rec7.phases.phase1.attachment_saved = True

            # Record 8: BAD - Processing (Forward Prepared)
            rec8 = Record(
                id="DEMO_REC_8",
                entry_id="ENTRY_8",
                center="EMS-B",
                type="BAD",
                ship_out_date=(datetime.date.today() - datetime.timedelta(days=1)).strftime("202X/%m/%d"),
                qty=12,
                subject="[APAC] Repair 坏品_入Hub-G_20260521_12PCS",
                received_time="202X/05/21 09:45:--",
                attachments=["packing list-WEST-702.xlsx", "shipping invoice-WEST-702.xlsx", "shipping list-12PCS不良品.xlsx"],
                keyword="WEST-702",
                batch="WEST-702",
                transfer_type="real",
                calculated_subject="Repair_BAD_return to hub_real_Hub_20260521",
                status="Processing"
            )
            rec8.phases.phase1.forward_prepared = True

            # Record 9: OK - New
            rec9 = Record(
                id="DEMO_REC_9",
                entry_id="ENTRY_9",
                center="EMS-C",
                type="OK",
                ship_out_date=today_str,
                qty=150,
                subject=f"[APAC] Repair 好品_入Hub-G_{today_nodash}_150PCS",
                received_time=f"{today_str} 10:15:--",
                attachments=["packing list-DOA-901.xlsx", "shipping invoice-DOA-901.xlsx"],
                keyword="DOA-901",
                batch="DOA-901",
                transfer_type="real",
                calculated_subject=f"Repair_OK_return to hub_real_Hub-G_{today_nodash}",
                status="New"
            )

                        # Record 10: OK - Shipped (Historical Date: 45 days ago)
            rec10 = Record(
                id="DEMO_REC_10",
                entry_id="ENTRY_10",
                center="EMS-A",
                type="OK",
                ship_out_date=(datetime.date.today() - datetime.timedelta(days=45)).strftime("202X/%m/%d"),
                qty=80,
                creation_date=(datetime.date.today() - datetime.timedelta(days=44)).strftime("202X/%m/%d"),
                er="TRK-XX-11010",
                subject="[APAC] Repair 好品_入Hub-G_20260401_80PCS",
                received_time="202X/04/01 09:15:--",
                attachments=["packing list-HUB-410.xlsx", "shipping invoice-HUB-410.xlsx"],
                keyword="HUB-410",
                batch="HUB-410",
                transfer_type="virtual",
                calculated_subject="Repair_OK_return to hub_virtual_Hub-G_20260401",
                status="Shipped"
            )
            rec10.phases.phase1.completed = True
            rec10.phases.phase1.attachment_saved = True
            rec10.phases.phase1.forward_prepared = True
            rec10.phases.phase1.completed_time = "202X/04/01 10:30:--"
            rec10.phases.phase2.completed = True
            rec10.phases.phase2.available = True
            rec10.phases.phase2.shipped = True
            rec10.phases.phase2.ba_no = "TRK-XX-11010"
            rec10.phases.phase2.completed_time = "202X/04/02 11:20:--"

            # Record 11: BAD - RMA Approved (Historical Date: 60 days ago)
            rec11 = Record(
                id="DEMO_REC_11",
                entry_id="ENTRY_11",
                center="EMS-B",
                type="BAD",
                ship_out_date=(datetime.date.today() - datetime.timedelta(days=60)).strftime("202X/%m/%d"),
                qty=25,
                creation_date=(datetime.date.today() - datetime.timedelta(days=59)).strftime("202X/%m/%d"),
                er="TRK-XX-11011",
                rma_no="RMA-XXX-XXXX, RMA-YYY-YYYY",
                subject="[APAC] Repair 坏品_入Hub-G_20260315_25PCS",
                received_time="202X/03/15 10:45:--",
                attachments=["packing list-HUB-311.xlsx", "shipping invoice-HUB-311.xlsx", "shipping list-25PCS不良品.xlsx"],
                keyword="HUB-311",
                batch="HUB-311",
                transfer_type="virtual",
                calculated_subject="Repair_BAD_return to hub_virtual_Hub_20260315",
                status="Shipped"
            )
            rec11.phases.phase1.completed = True
            rec11.phases.phase1.attachment_saved = True
            rec11.phases.phase1.forward_prepared = True
            rec11.phases.phase1.completed_time = "202X/03/15 11:50:--"
            rec11.phases.phase2.completed = True
            rec11.phases.phase2.available = True
            rec11.phases.phase2.shipped = True
            rec11.phases.phase2.ba_no = "TRK-XX-11011"
            rec11.phases.phase2.completed_time = "202X/03/16 15:10:--"
            rec11.phases.phase3.completed = True
            rec11.phases.phase3.completed_time = "202X/03/17 09:30:--"

            db.records.extend([rec1, rec2, rec3, rec4, rec5, rec6, rec7, rec8, rec9, rec10, rec11])
            ship_logger.info("Demo Mode: Seeded 11 initial records covering all stages (including historical data)")
        else:
            # Dynamically simulate incoming system updates when user scans again!
            for r in db.records:
                if r.status == "RPA Submitted" and not r.phases.phase2.available:
                    r.phases.phase2.available = True
                    suffix = r.id[-1] if r.id[-1].isdigit() else "5"
                    r.phases.phase2.ba_no = f"TRK-XX-1100{suffix}"
                    r.phases.phase2.entry_id = f"DEMO_PHASE2_REPLY_{r.id}"
                    r.er = r.phases.phase2.ba_no
                    r.creation_date = datetime.date.today().strftime("202X/%m/%d")
                    r.current_phase = 2
                    r.status = "Phase 2 Ready"
                    ship_logger.info("Demo Mode: Simulated Phase 2 EMS reply with er=%s", r.er)
                elif r.status == "Shipped" and r.type == "BAD" and not r.phases.phase3.completed:
                    r.phases.phase3.completed = True
                    r.phases.phase3.completed_time = datetime.datetime.now().strftime("202X/%m/%d %H:%M:--")
                    r.rma_no = "RMA-XXX-XXXX, RMA-YYY-YYYY"
                    ship_logger.info("Demo Mode: Simulated Phase 3 RMA approval for %s", r.id)
                elif r.status == "New":
                    r.phases.phase1.attachment_saved = True
                    r.status = "Processing"
                    ship_logger.info("Demo Mode: Auto-progressed New to Processing for %s", r.id)
        
        for record in db.records:
            if record.phases.phase2.completed or record.phases.phase2.shipped:
                record.status = "Shipped"
            elif record.phases.phase2.available:
                record.status = "Phase 2 Ready"
            elif record.phases.phase1.completed:
                record.status = "RPA Submitted"
            elif record.phases.phase1.attachment_saved or record.phases.phase1.forward_prepared:
                record.status = "Processing"
            else:
                record.status = "New"

        save_db(db)
        return db.records

    # The original live Outlook scan path (never reached unless check_outlook_available() is forced True)
    new_records_count = 0
    with OutlookCOMContext() as outlook:
        try:
            ns = outlook.GetNamespace("MAPI")
            inbox = ns.GetDefaultFolder(6)
            messages = inbox.Items
            messages.Sort("[ReceivedTime]", True)
            
            limit = min(60, messages.Count)
            for i in range(1, limit + 1):
                try:
                    msg = messages.Item(i)
                    subject = str(msg.Subject or "")
                    sender = ""
                    sender_name = ""
                    try:
                        sender = str(msg.SenderEmailAddress or "").lower()
                        if "/o=" in sender or "@" not in sender:
                            try:
                                sender_obj = msg.Sender
                                if sender_obj:
                                    ex_user = sender_obj.GetExchangeUser()
                                    if ex_user:
                                        smtp_addr = ex_user.PrimarySmtpAddress
                                        if smtp_addr:
                                            sender = smtp_addr.lower()
                            except:
                                pass
                    except:
                        sender = ""

                    try:
                        sender_name = str(msg.SenderName or "").lower()
                    except:
                        sender_name = ""

                    is_sarah = "sarah" in sender or "sarah" in sender_name
                    phase3_info = extract_phase3_info(subject) if is_sarah else None
                    if phase3_info:
                        matched = False
                        for record in db.records:
                            if record.type == "BAD" and record.er:
                                ems_clean = record.er.replace("BA-", "").strip().upper()
                                email_ba_clean = phase3_info["ba_no"].replace("BA-", "").strip().upper()
                                if ems_clean and ems_clean == email_ba_clean:
                                    body_text = getattr(msg, "Body", "")
                                    unique_ides = []
                                    for ide in re.findall(r"IDE[a-zA-Z0-9]+", body_text):
                                        ide_upper = ide.upper()
                                        if ide_upper not in unique_ides:
                                            unique_ides.append(ide_upper)
                                    unique_ides = unique_ides[:4]
                                    if unique_ides:
                                        masked_list = ["RMA-XXX-XXXX", "RMA-YYY-YYYY", "RMA-ZZZ-ZZZZ", "RMA-AAA-AAAA"]
                                        record.rma_no = ", ".join(masked_list[:len(unique_ides)])
                                        record.phases.phase3.completed = True
                                        try:
                                            record.phases.phase3.completed_time = msg.ReceivedTime.strftime("202X/%m/%d %H:%M:--")
                                        except:
                                            record.phases.phase3.completed_time = datetime.datetime.now().strftime("202X/%m/%d %H:%M:--")
                                        matched = True
                                        break
                        if matched:
                            continue

                    phase2_info = extract_phase2_info(subject)
                    if phase2_info:
                        matched = False
                        for record in db.records:
                            if record_matches_phase2(record, phase2_info):
                                record.phases.phase2.available = True
                                record.phases.phase2.ba_no = phase2_info["ba_no"]
                                record.phases.phase2.entry_id = str(msg.EntryID)
                                ba_clean = phase2_info['ba_no'].replace("BA-", "").replace("TRK-XX-", "")
                                record.phases.phase2.ba_no = f"TRK-XX-{ba_clean}"
                                record.er = record.phases.phase2.ba_no
                                
                                try:
                                    record.creation_date = msg.ReceivedTime.strftime("202X/%m/%d")
                                except:
                                    record.creation_date = datetime.datetime.now().strftime("202X/%m/%d")
                                    
                                record.current_phase = 2
                                record.status = "Phase 2 Ready"
                                matched = True
                                break
                        continue

                    parsed = parse_email_subject(subject)
                    if parsed:
                        entry_id = str(msg.EntryID)
                        record_id = entry_id[-20:] if len(entry_id) > 20 else entry_id
                        
                        existing = next((r for r in db.records if r.entry_id == entry_id or r.id == record_id), None)
                        if not existing:
                            attachments_list = []
                            keyword = ""
                            for j in range(1, msg.Attachments.Count + 1):
                                att = msg.Attachments.Item(j)
                                filename = str(att.FileName)
                                attachments_list.append(filename)
                                if filename.lower().startswith("packing list-") and not keyword:
                                    keyword = extract_keyword_from_attachment(filename)
                            
                            if not keyword:
                                keyword = "Unknown"

                            batch = extract_batch_from_attachments(attachments_list)
                            today_str = datetime.date.today().strftime("202X%m%d")
                            trans_details = calculate_transfer_and_subject(parsed["type"], keyword, today_str)
                            
                            try:
                                received_time = msg.ReceivedTime.strftime("202X/%m/%d %H:%M:--")
                            except:
                                received_time = datetime.datetime.now().strftime("202X/%m/%d %H:%M:--")
                            
                            new_record = Record(
                                id=record_id,
                                entry_id=entry_id,
                                center=parsed["center"],
                                type=parsed["type"],
                                ship_out_date=parsed["ship_out_date"],
                                qty=parsed["qty"],
                                subject=subject,
                                received_time=received_time,
                                attachments=attachments_list,
                                keyword=keyword,
                                batch=batch,
                                transfer_type=trans_details["transfer_type"],
                                calculated_subject=trans_details["calculated_subject"]
                            )
                            db.records.append(new_record)
                            new_records_count += 1
                except Exception as ex:
                    print(f"Skipping individual email due to read error: {ex}")
                    continue
                    
        except Exception as e:
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Outlook reading error: {str(e)}")
            
    for record in db.records:
        if record.phases.phase2.completed or record.phases.phase2.shipped:
            record.status = "Shipped"
        elif record.phases.phase2.available:
            record.status = "Phase 2 Ready"
        elif record.phases.phase1.completed:
            record.status = "RPA Submitted"
        elif record.phases.phase1.attachment_saved or record.phases.phase1.forward_prepared:
            record.status = "Processing"
        else:
            record.status = "New"

    save_db(db)
    return db.records

@app.put("/api/records/{record_id}", response_model=Record)
def update_record_fields(record_id: str, payload: Dict = Body(...)):
    db = load_db()
    record = next((r for r in db.records if r.id == record_id), None)
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
        
    if "creation_date" in payload:
        record.creation_date = str(payload["creation_date"])
    if "bpm_date" in payload:
        record.creation_date = str(payload["bpm_date"])
    if "er" in payload:
        record.er = str(payload["er"])
    if "ems_no" in payload:
        record.er = str(payload["ems_no"])
    if "bpm_no" in payload:
        record.er = str(payload["bpm_no"])
    if "rma_no" in payload:
        record.rma_no = str(payload["rma_no"])
    if "return_to" in payload:
        record.return_to = str(payload["return_to"])
    if "center" in payload:
        record.center = str(payload["center"])
    if "rtr" in payload:
        record.center = str(payload["rtr"])
    if "phase1_completed_time" in payload:
        record.phases.phase1.completed_time = str(payload["phase1_completed_time"])
        record.phases.phase1.completed = True
        
    save_db(db)
    return record

@app.post("/api/records/{record_id}/save_attachments", response_model=Record)
def save_attachments_endpoint(record_id: str):
    db = load_db()
    record = next((r for r in db.records if r.id == record_id), None)
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
        
    try:
        ship_date = datetime.datetime.strptime(record.ship_out_date, "%Y/%m/%d").date()
    except Exception:
        ship_date = datetime.date.today()
    week_str = get_week_number(ship_date)
    
    batch_value = record.batch if record.batch else (record.keyword if record.keyword else "Unknown")
    main_folder_name = f"{week_str}[{batch_value}]"
    sub_folder_name = record.type  # "OK" or "BAD"
    
    dest_dir = os.path.join(db.config.download_dir, main_folder_name, sub_folder_name)
    os.makedirs(dest_dir, exist_ok=True)
    
    # ---------------- DEMO MODE FALLBACK ----------------
    if not check_outlook_available() or record.id.startswith("DEMO_"):
        ship_logger.info("Running save_attachments in [DEMO MODE] fallback")
        if not record.attachments:
            if record.type == "OK":
                record.attachments = [f"packing list-{batch_value}.xlsx", f"shipping invoice-{batch_value}.xlsx"]
            else:
                record.attachments = [f"packing list-{batch_value}.xlsx", f"shipping invoice-{batch_value}.xlsx", f"shipping list-{record.qty}PCS不良品.xlsx"]
        
        create_mock_attachment_files(dest_dir, record.attachments)
        
        record.phases.phase1.attachment_saved = True
        if record.phases.phase2.shipped or record.phases.phase2.completed:
            record.status = "Shipped"
        elif record.phases.phase1.forward_prepared:
            record.phases.phase1.completed = True
            record.phases.phase1.completed_time = datetime.datetime.now().strftime("202X/%m/%d %H:%M:--")
            record.status = "RPA Submitted"
        else:
            record.status = "Processing"
            
        save_db(db)
        return record
    # ----------------------------------------------------

    with OutlookCOMContext() as outlook:
        try:
            ns = outlook.GetNamespace("MAPI")
            msg = ns.GetItemFromID(record.entry_id)
            
            if msg.Attachments.Count == 0:
                raise HTTPException(status_code=400, detail="This email has no attachments to save.")
            
            skip_extensions = {'.png', '.jpg', '.jpeg', '.gif'}
            for i in range(1, msg.Attachments.Count + 1):
                att = msg.Attachments.Item(i)
                filename = str(att.FileName)
                file_ext = os.path.splitext(filename)[1].lower()
                if file_ext in skip_extensions:
                    continue
                file_path = os.path.join(dest_dir, filename)
                att.SaveAsFile(file_path)
                
        except Exception as e:
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Failed to save attachments via Outlook COM: {str(e)}")
            
    record.phases.phase1.attachment_saved = True
    if record.phases.phase2.shipped or record.phases.phase2.completed:
        record.status = "Shipped"
    elif record.phases.phase1.forward_prepared:
        record.phases.phase1.completed = True
        record.phases.phase1.completed_time = datetime.datetime.now().strftime("202X/%m/%d %H:%M:--")
        record.status = "RPA Submitted"
    else:
        record.status = "Processing"
        
    save_db(db)
    return record

@app.get("/api/export_excel")
def export_excel_endpoint():
    return export_excel()

@app.post("/api/records/{record_id}/ship", response_model=Record)
def ship_endpoint(record_id: str):
    ship_logger.info("ship_endpoint invoked: record_id=%s", record_id)
    db = load_db()
    record = next((r for r in db.records if r.id == record_id), None)
    if not record:
        ship_logger.info("ship_endpoint missing record: record_id=%s", record_id)
        raise HTTPException(status_code=404, detail="Record not found")
    
    # Ensure EMS No is present
    if not record.phases.phase2.ba_no:
        suffix = record.id[-1] if record.id[-1].isdigit() else "5"
        record.phases.phase2.ba_no = f"TRK-XX-1100{suffix}"
        record.er = record.phases.phase2.ba_no
        record.creation_date = datetime.date.today().strftime("202X/%m/%d")
        record.current_phase = 2
        record.phases.phase2.available = True
    
    # ---------------- DEMO MODE FALLBACK ----------------
    if not check_outlook_available() or record.id.startswith("DEMO_"):
        ship_logger.info("Running ship_endpoint in [DEMO MODE] fallback")
        
        ba_no = record.phases.phase2.ba_no
        if ba_no:
            ba_no = f"TRK-XX-{ba_no.replace('BA-', '').replace('TRK-XX-', '')}"
        qty = record.qty
        
        if record.type == "OK":
            to_addr = "alice.thompson@client.example"
            cc_addr = (
                "tom.harvey@logistics.example; "
                "wh.logistics@logistics.example; "
                "sh.wh1@logistics.example; "
                "alex.mercer@logistics.example; "
                "fiona.gallagher@logistics.example; "
                "iqc.team@logistics.example; "
                "brandon.wu@logistics.example; "
                "derek.stone@logistics.example; "
                "victor.vance@logistics.example"
            )
            html_body = f"""<div style="font-size: 11pt; font-family: 'Calibri', sans-serif; line-height: 1.5;">
Hi Alice,<br><br>
OK items {qty}pcs, Tracking No: {ba_no}, please arrange. Thank you!<br>
</div>"""
        else:
            to_addr = "robert.chen@client.example; sarah.jenkins@client.example"
            cc_addr = (
                "tom.harvey@logistics.example; "
                "wh.logistics@logistics.example; "
                "sh.wh1@logistics.example; "
                "sh.wh2@logistics.example; "
                "alex.mercer@logistics.example; "
                "fiona.gallagher@logistics.example; "
                "iqc.team@logistics.example; "
                "brandon.wu@logistics.example; "
                "derek.stone@logistics.example; "
                "victor.vance@logistics.example; "
                "david.walsh@client.example; "
                "emily.rodriguez@client.example; "
                "marcus.vance@client.example; "
                "grace.lee@client.example; "
                "lucas.silva@client.example; "
                "kevin.parker@client.example; "
                "sophia.martinez@client.example; "
                "chloe.harrison@client.example; "
                "daniel.kim@client.example"
            )
            html_body = f"""<div style="font-size: 11pt; font-family: 'Calibri', sans-serif; line-height: 1.5;">
Hi Robert,<br>
Tracking No: {ba_no}<br>
Please arrange the return of {qty}pcs dead boards to Hub. Thanks!<br><br>
Hi Sarah,<br>
Please find attached {qty}pcs for the RMA application. Thanks!<br>
</div>"""
            
        generate_simulated_outlook_window(
            subject=f"RE: Repair_{record.type}_return to hub_{record.transfer_type}_Hub-G_{record.received_time.split(' ')[0].replace('/', '')}",
            to_address=to_addr,
            cc_address=cc_addr,
            body=html_body,
            attachments=record.attachments,
            filename_prefix=f"Simulated_Ship_Reply_{record.id}"
        )
        
        record.phases.phase2.shipped = True
        record.phases.phase2.completed = True
        record.phases.phase2.completed_time = datetime.datetime.now().strftime("202X/%m/%d %H:%M:--")
        record.status = "Shipped"
        
        save_db(db)
        return record
    # ----------------------------------------------------

    with OutlookCOMContext() as outlook:
        try:
            ns = outlook.GetNamespace("MAPI")
            msg = None
            if record.phases.phase2.entry_id:
                try:
                    msg = ns.GetItemFromID(record.phases.phase2.entry_id)
                except Exception as ex:
                    ship_logger.info("GetItemFromID failed for phase2 entry_id=%s: %s. Will re-search inbox.", record.phases.phase2.entry_id, str(ex))
                    msg = None

            if not msg:
                ship_logger.info("ship_endpoint searching inbox for phase2 reply: record_id=%s", record_id)
                phase2_match = find_phase2_reply_for_record(record, ns)
                if phase2_match:
                    info = phase2_match["info"]
                    record.phases.phase2.entry_id = phase2_match["entry_id"]
                    record.phases.phase2.ba_no = info["ba_no"]
                    record.phases.phase2.available = True
                    ba_clean = info['ba_no'].replace("BA-", "").replace("TRK-XX-", "")
                    record.phases.phase2.ba_no = f"TRK-XX-{ba_clean}"
                    record.er = record.phases.phase2.ba_no
                    record.current_phase = 2
                    record.status = "Phase 2 Ready"
                    try:
                        if phase2_match["received_time"]:
                            record.creation_date = phase2_match["received_time"].strftime("202X/%m/%d")
                    except Exception:
                        pass
                    save_db(db)
                    
                    try:
                        msg = ns.GetItemFromID(record.phases.phase2.entry_id)
                    except Exception as ex:
                        ship_logger.error("Failed to load freshly matched phase2 message: %s", str(ex))
                        msg = None
                else:
                    ship_logger.info("ship_endpoint no phase2 reply found: record_id=%s", record_id)
                    raise HTTPException(status_code=400, detail="未找到對應的第二階段郵件 (可能已被搬移或刪除)，請先確認收到回覆信件。")

            if not msg:
                raise HTTPException(status_code=400, detail="無法讀取第二階段郵件對象，請確認該郵件在 Outlook 中是否正常。")
            
            reply_msg = msg.ReplyAll()
            
            # Find and collect attachments
            unique_attachments = {}
            temp_dir = tempfile.mkdtemp()

            def collect_attachments_from_msg(message, source_label):
                for i in range(1, message.Attachments.Count + 1):
                    att = message.Attachments.Item(i)
                    fname = str(att.FileName)
                    lname = fname.lower()
                    matched = False
                    for attachment_type, keywords in ATTACHMENT_TYPE_KEYWORDS.items():
                        for keyword in keywords:
                            if keyword in lname:
                                fpath = os.path.join(temp_dir, fname)
                                att.SaveAsFile(fpath)
                                unique_attachments[fname.lower()] = fpath
                                matched = True
                                break
                        if matched:
                            break

            try:
                orig_msg = ns.GetItemFromID(record.entry_id)
                collect_attachments_from_msg(orig_msg, "original email")
            except Exception as ex:
                ship_logger.info("Original email attachment scan failed: %s", str(ex))

            try:
                phase2_msg = ns.GetItemFromID(record.phases.phase2.entry_id)
                collect_attachments_from_msg(phase2_msg, "phase2 email")
            except Exception as ex:
                ship_logger.info("Phase2 email attachment scan failed: %s", str(ex))

            # Search in storage
            storage_dirs = []
            search_date = None
            if record.phases.phase1.completed_time:
                try:
                    dt_str = record.phases.phase1.completed_time.split(" ")[0]
                    search_date = datetime.datetime.strptime(dt_str, "%Y/%m/%d").date()
                except Exception:
                    search_date = None
            if not search_date:
                try:
                    search_date = datetime.datetime.strptime(record.received_time.split(" ")[0], "%Y/%m/%d").date()
                except Exception:
                    search_date = None
            if not search_date:
                try:
                    parts = record.ship_out_date.replace("/", "").strip()
                    if len(parts) == 8:
                        search_date = datetime.datetime.strptime(parts, "%Y%m%d").date()
                except Exception:
                    search_date = None
            if not search_date:
                search_date = datetime.date.today()

            week_str = get_week_number(search_date)
            batch_val = record.batch if record.batch else (record.keyword if record.keyword else "Unknown")
            folder_name = f"{week_str}[{batch_val}]"
            guess_dir = os.path.join(db.config.download_dir, folder_name, record.type)
            storage_dirs.append(guess_dir)

            if not os.path.exists(guess_dir):
                for root, dirs, files in os.walk(db.config.download_dir):
                    if os.path.basename(root).lower() == record.type.lower():
                        storage_dirs.append(root)

            batch_tokens = []
            if record.batch:
                batch_tokens = [t.strip() for t in record.batch.split("&") if t.strip()]
            elif record.keyword:
                batch_tokens = [record.keyword.strip()]

            for dest_dir in storage_dirs:
                if not os.path.exists(dest_dir):
                    continue
                for f in os.listdir(dest_dir):
                    fname_lower = f.lower()
                    is_packing = any(kw in fname_lower for kw in ATTACHMENT_TYPE_KEYWORDS["packing list"])
                    is_shipping = any(kw in fname_lower for kw in ATTACHMENT_TYPE_KEYWORDS["shipping list"])
                    
                    if is_packing or is_shipping:
                        matched_batch = False
                        if batch_tokens:
                            for token in batch_tokens:
                                if token.lower() in fname_lower:
                                    matched_batch = True
                                    break
                        else:
                            matched_batch = True
                            
                        if matched_batch:
                            fpath = os.path.join(dest_dir, f)
                            if fname_lower not in unique_attachments:
                                unique_attachments[fname_lower] = fpath

            found_packing = has_type(list(unique_attachments.values()), "packing list")
            found_shipping = has_type(list(unique_attachments.values()), "shipping list")

            if not found_packing and not found_shipping:
                shutil.rmtree(temp_dir, ignore_errors=True)
                raise HTTPException(status_code=400, detail="目前沒有 packing list 或 shipping list，無法完成 Ship 動作。")

            # Attach files
            for fpath in unique_attachments.values():
                reply_msg.Attachments.Add(fpath)

            ba_no = record.phases.phase2.ba_no
            if ba_no:
                ba_no = f"TRK-XX-{ba_no.replace('BA-', '').replace('TRK-XX-', '')}"
            qty = record.qty
            
            if record.type == "OK":
                reply_msg.To = "alice.thompson@client.example"
                reply_msg.CC = (
                    "tom.harvey@logistics.example; "
                    "wh.logistics@logistics.example; "
                    "sh.wh1@logistics.example; "
                    "alex.mercer@logistics.example; "
                    "fiona.gallagher@logistics.example; "
                    "iqc.team@logistics.example; "
                    "brandon.wu@logistics.example; "
                    "derek.stone@logistics.example; "
                    "victor.vance@logistics.example"
                )
                html_body = f"""<div style="font-size: 11pt; font-family: 'Calibri', sans-serif; line-height: 1.5;">
Hi Alice,<br><br>
OK items {qty}pcs, Tracking No: {ba_no}, please arrange. Thank you!<br>
</div>"""
                reply_msg.HTMLBody = html_body + reply_msg.HTMLBody
            else:
                reply_msg.To = "robert.chen@client.example; sarah.jenkins@client.example"
                reply_msg.CC = (
                    "tom.harvey@logistics.example; "
                    "wh.logistics@logistics.example; "
                    "sh.wh1@logistics.example; "
                    "sh.wh2@logistics.example; "
                    "alex.mercer@logistics.example; "
                    "fiona.gallagher@logistics.example; "
                    "iqc.team@logistics.example; "
                    "brandon.wu@logistics.example; "
                    "derek.stone@logistics.example; "
                    "victor.vance@logistics.example; "
                    "david.walsh@client.example; "
                    "emily.rodriguez@client.example; "
                    "marcus.vance@client.example; "
                    "grace.lee@client.example; "
                    "lucas.silva@client.example; "
                    "kevin.parker@client.example; "
                    "sophia.martinez@client.example; "
                    "chloe.harrison@client.example; "
                    "daniel.kim@client.example"
                )
                html_body = f"""<div style="font-size: 11pt; font-family: 'Calibri', sans-serif; line-height: 1.5;">
Hi Robert,<br>
Tracking No: {ba_no}<br>
Please arrange the return of {qty}pcs dead boards to Hub. Thanks!<br><br>
Hi Sarah,<br>
Please find attached {qty}pcs for the RMA application. Thanks!<br>
</div>"""
                reply_msg.HTMLBody = html_body + reply_msg.HTMLBody

            try:
                reply_msg.Recipients.ResolveAll()
            except Exception as rex:
                ship_logger.info("Recipients ResolveAll failed for reply_msg: %s", str(rex))

            reply_msg.Display()
            shutil.rmtree(temp_dir, ignore_errors=True)

        except HTTPException:
            raise
        except Exception as e:
            ship_logger.error("ship_endpoint exception for record_id=%s: %s", record_id, traceback.format_exc())
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Failed to prepare ship reply: {str(e)}")
            
    record.phases.phase2.shipped = True
    record.phases.phase2.completed = True
    record.phases.phase2.completed_time = datetime.datetime.now().strftime("202X/%m/%d %H:%M:--")
    record.status = "Shipped"
    
    save_db(db)
    return record

@app.post("/api/records/{record_id}/prepare_forward", response_model=Record)
def prepare_forward_endpoint(record_id: str):
    db = load_db()
    record = next((r for r in db.records if r.id == record_id), None)
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
        
    today_str = datetime.date.today().strftime("202X%m%d")
    trans_details = calculate_transfer_and_subject(record.type, record.keyword, today_str)
    
    record.calculated_subject = trans_details["calculated_subject"]
    record.transfer_type = trans_details["transfer_type"]

    # ---------------- DEMO MODE FALLBACK ----------------
    if not check_outlook_available() or record.id.startswith("DEMO_"):
        ship_logger.info("Running prepare_forward in [DEMO MODE] fallback")
        
        html_prefix = """<div style="font-size: 11pt; line-height: 1.5; font-family: 'Calibri', sans-serif;">
Hi Team,<br>
Please find the attached shipping data. Please assist with submitting the inventory value and issuing the transfer slip, and kindly provide the Tracking No. Thank you!<br>
</div>"""
        
        generate_simulated_outlook_window(
            subject=record.calculated_subject,
            to_address="rpa-bom@client.example",
            cc_address="oliver.smith@client.example",
            body=html_prefix,
            attachments=record.attachments,
            filename_prefix=f"Simulated_RPA_Forward_Draft_{record.id}"
        )
        
        record.phases.phase1.forward_prepared = True
        if record.phases.phase2.shipped or record.phases.phase2.completed:
            record.status = "Shipped"
        elif record.phases.phase1.attachment_saved:
            record.phases.phase1.completed = True
            record.phases.phase1.completed_time = datetime.datetime.now().strftime("202X/%m/%d %H:%M:--")
            record.status = "RPA Submitted"
        else:
            record.status = "Processing"
            
        save_db(db)
        return record
    # ----------------------------------------------------

    with OutlookCOMContext() as outlook:
        try:
            ns = outlook.GetNamespace("MAPI")
            msg = ns.GetItemFromID(record.entry_id)
            
            forward_msg = msg.Forward()
            forward_msg.Subject = record.calculated_subject
            forward_msg.To = "rpa-bom@client.example"
            forward_msg.CC = "oliver.smith@client.example"
            
            html_prefix = """<div style="font-size: 11pt; line-height: 1.5; font-family: 'Calibri', sans-serif;">
Hi Team,<br>
Please find the attached shipping data. Please assist with submitting the inventory value and issuing the transfer slip, and kindly provide the Tracking No. Thank you!<br>
</div>"""
            forward_msg.HTMLBody = html_prefix + forward_msg.HTMLBody
            
            try:
                forward_msg.Recipients.ResolveAll()
            except Exception as rex:
                print(f"Recipients ResolveAll failed for forward_msg: {rex}")

            forward_msg.Display()
            
        except Exception as e:
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Failed to prepare Outlook forward draft: {str(e)}")
            
    record.phases.phase1.forward_prepared = True
    if record.phases.phase2.shipped or record.phases.phase2.completed:
        record.status = "Shipped"
    elif record.phases.phase1.attachment_saved:
        record.phases.phase1.completed = True
        record.phases.phase1.completed_time = datetime.datetime.now().strftime("202X/%m/%d %H:%M:--")
        record.status = "RPA Submitted"
    else:
        record.status = "Processing"
        
    save_db(db)
    return record

import tkinter as tk
from tkinter import filedialog

@app.get("/api/select_folder")
def select_folder():
    try:
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        folder_path = filedialog.askdirectory(title="Select Attachment Storage Folder")
        root.destroy()
        return {"folder_path": folder_path}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Serve frontend HTML
@app.get("/", response_class=HTMLResponse)
def index():
    html_file = os.path.join(TEMPLATES_DIR, "index.html")
    if not os.path.exists(html_file):
        raise HTTPException(status_code=404, detail="Frontend index.html not found.")
    with open(html_file, "r", encoding="utf-8") as f:
        return f.read()

if __name__ == "__main__":
    import webbrowser
    import threading
    import time
    
    def open_browser():
        time.sleep(1.5)  # Wait for uvicorn to boot up
        webbrowser.open("http://127.0.0.1:8000")
        
    threading.Thread(target=open_browser, daemon=True).start()

    # Start FastAPI server on port 8000
    uvicorn.run("app:app", host="127.0.0.1", port=8000, reload=True)
